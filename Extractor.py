import urllib.request
import urllib.parse
import http.cookiejar
import json
from lxml import etree
import os
import concurrent.futures
import re
import sqlite3
import codecs
import shutil
import os.path


def dict_factory(cursor, row):
    d = {}
    for idx, col in enumerate(cursor.description):
        d[col[0]] = row[idx]
    return d


class Extractor:
    def __init__(self):
        self.base_url = 'https://leetcode.com'
        self.db_name = 'leetcode.db'
        cj = http.cookiejar.CookieJar()
        self.opener = urllib.request.build_opener(urllib.request.HTTPCookieProcessor(cj))
        self.opener.addheaders = [
            ('Host', 'leetcode.com'),
            ('User-Agent',
             'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_12_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36')
        ]
        self.is_logged_in = False

    def login(self, user_name, password):
        if self.is_logged_in:
            return
        url = self.base_url + '/accounts/login/'
        with self.opener.open(url) as f:
            content = f.read().decode('utf-8')
        token = re.findall("name='csrfmiddlewaretoken'\svalue='(.*?)'", content)[0]
        post_data = {
            'csrfmiddlewaretoken': token,
            'login': user_name,
            'password': password
        }
        post_data = urllib.parse.urlencode(post_data)
        self.opener.addheaders.append(('Referer', url))
        with self.opener.open(url, data=post_data.encode()) as f:
            if f.read().decode().find('Successfully signed in') != -1:
                self.is_logged_in = True
                print('logged in')
            else:
                print('failed to login in')
        self.opener.addheaders.pop()

    def get_problem_list(self):
        with self.opener.open(self.base_url + '/api/problems/algorithms/') as f:
            content = f.read().decode('utf-8')
        content = json.loads(content)
        return content['stat_status_pairs']

    def store_problem_list_to_db(self, problem_list):
        conn = sqlite3.connect(self.db_name)
        c = conn.cursor()
        c.execute(
            '''
                CREATE TABLE IF NOT EXISTS problem (
                    id INTEGER,
                    title TEXT,
                    slug TEXT,
                    difficulty INTEGER,
                    paid_only INTEGER,
                    status TEXT,
                    total_acs INTEGER,
                    total_submitted INTEGER,
                    PRIMARY KEY(id))
            ''')
        c.execute('DELETE FROM problem')
        for problem in problem_list:
            c.execute(
                '''
                    INSERT INTO problem 
                        (id, title, slug, difficulty, paid_only, status, total_acs, total_submitted) 
                    VALUES 
                        (?, ?, ?, ?, ?, ?, ?, ?)
                ''',
                (problem['stat']['question_id']
                 , problem['stat']['question__title']
                 , problem['stat']['question__title_slug']
                 , problem['difficulty']['level']
                 , 1 if problem['paid_only'] else 0
                 , problem['status']
                 , problem['stat']['total_acs']
                 , problem['stat']['total_submitted'])
            )
        conn.commit()
        conn.close()

    def update_problem_list(self):
        self.store_problem_list_to_db(self.get_problem_list())

    def get_problem_list_from_db(self):
        conn = sqlite3.connect(self.db_name)
        conn.row_factory = dict_factory
        c = conn.cursor()
        c.execute('SELECT * FROM problem')
        return c.fetchall()

    def get_description(self, url, file_path):
        with self.opener.open(url) as f:
            content = f.read().decode('utf-8')
            root = etree.HTML(content)
            result = root.xpath('//*[@id="descriptionContent"]//div[@class="question-description"]')
            html = etree.tostring(result[0], encoding='utf-8')
            with open(file_path, 'wb') as f:
                f.write(html)
        return file_path

    def extract_descriptions(self):
        conn = sqlite3.connect(self.db_name)
        c = conn.cursor()
        c.execute('CREATE TABLE IF NOT EXISTS description (title TEXT, path TEXT, PRIMARY KEY(title))')
        c.execute(
            '''
                SELECT a.id, a.title, a.slug 
                FROM problem a 
                LEFT JOIN description b 
                ON a.title=b.title 
                WHERE a.paid_only=0 AND b.title IS NULL
            ''')
        problems = c.fetchall()
        dir_path = 'descriptions/'
        os.makedirs(dir_path, exist_ok=True)
        with concurrent.futures.ThreadPoolExecutor(max_workers=20) as executor:
            futures = {
                executor.submit(self.get_description
                                , self.base_url + '/problems/' + problem[2] + '/description/'
                                , os.path.join(dir_path, str(problem[0]).zfill(3) + '. ' + problem[1] + '.html')):
                    problem[1] for problem in problems}
            for future in concurrent.futures.as_completed(futures):
                title = futures[future]
                try:
                    file_path = future.result()
                except Exception as e:
                    print('%r generated an exception: %s' % (title, e))
                else:
                    if file_path:
                        c.execute('INSERT INTO description (title, path) VALUES (?, ?)', (title, file_path))
        conn.commit()
        conn.close()

    def sync_description_db_and_file(self):
        conn = sqlite3.connect(self.db_name)
        conn.row_factory = dict_factory
        c = conn.cursor()
        c.execute('SELECT name FROM sqlite_master WHERE type="table" AND name="description"')
        if not c.fetchone():
            return
        c.execute('SELECT title,path FROM description')
        descriptions = c.fetchall()
        for description in descriptions:
            if not os.path.exists(description['path']):
                c.execute('DELETE FROM description WHERE title=?', (description['title'],))
        conn.commit()
        conn.close()

    def update_descriptions(self):
        self.sync_description_db_and_file()
        self.extract_descriptions()

    def get_submission_list(self):
        if not self.is_logged_in:
            print('should login first')
            return
        result = []
        offset = 0
        LIMIT = 100
        while True:
            url = self.base_url + '/api/submissions/?offset=' + str(offset) + '&limit=' + str(LIMIT)
            with self.opener.open(url) as f:
                content = f.read().decode('utf-8')
            content = json.loads(content)
            result.extend(content['submissions_dump'])
            if not content['has_next']:
                return result
            offset += LIMIT

    def store_submission_list_to_db(self, submission_list):
        conn = sqlite3.connect(self.db_name)
        c = conn.cursor()
        c.execute(
            '''
                CREATE TABLE IF NOT EXISTS submission (
                    lang TEXT,
                    title TEXT,
                    url TEXT,
                    downloaded INTEGER DEFAULT 0,
                    path TEXT,
                    removed INTEGER DEFAULT 0,
                    PRIMARY KEY(url))
            ''')
        for submission in submission_list:
            if submission['status_display'] == 'Accepted':
                c.execute('INSERT OR IGNORE INTO submission (lang, title, url) VALUES (?, ?, ?)'
                          , (submission['lang'], submission['title'], submission['url']))
        conn.commit()
        conn.close()

    def update_submission_list(self):
        self.store_submission_list_to_db(self.get_submission_list())

    def get_submission(self, url, file_path):
        with self.opener.open(url) as f:
            content = f.read().decode('utf-8')
            code = re.findall("submissionCode:\s'(.*?)',", content)[0]
            code = codecs.decode(code, 'unicode-escape')
            code = code.replace('\r\n', '\n')
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(code)
        return file_path

    def extract_submissions(self):
        conn = sqlite3.connect(self.db_name)
        c = conn.cursor()
        c.execute('SELECT url FROM submission WHERE downloaded=0 AND removed=0')
        urls = c.fetchall()
        urls = [url[0] for url in urls]
        dir_path = 'submissions/'
        os.makedirs(dir_path, exist_ok=True)
        with concurrent.futures.ThreadPoolExecutor(max_workers=20) as executor:
            futures = {executor.submit(self.get_submission, self.base_url + url,
                                       os.path.join(dir_path, url.split('/')[-2])): url for url in urls}
            for future in concurrent.futures.as_completed(futures):
                url = futures[future]
                try:
                    file_path = future.result()
                except Exception as e:
                    print('%r generated an exception: %s' % (url, e))
                else:
                    if file_path:
                        c.execute('UPDATE submission SET downloaded=1,path=? WHERE url=?', (file_path, url))
        conn.commit()
        conn.close()

    def sync_submission_db_and_file(self):
        conn = sqlite3.connect(self.db_name)
        conn.row_factory = dict_factory
        c = conn.cursor()
        c.execute('SELECT name FROM sqlite_master WHERE type="table" AND name="submission"')
        if not c.fetchone():
            return
        c.execute('SELECT url,path FROM submission')
        submissions = c.fetchall()
        for submission in submissions:
            if not os.path.exists(submission['path']):
                c.execute('DELETE FROM submission WHERE url=?', (submission['url'],))
        conn.commit()
        conn.close()

    def update_submissions(self):
        self.sync_submission_db_and_file()
        self.update_submission_list()
        self.extract_submissions()

    def output_submissions(self, dir_path='out_submissions/', latest_only=True):
        def lang_to_language(lang):
            if lang == 'python' or lang == 'python3':
                return 'Python'
            if lang == 'java':
                return 'Java'
            if lang == 'cpp':
                return 'C++'

        def lang_to_extension(lang):
            if lang == 'python' or lang == 'python3':
                return '.py'
            if lang == 'java':
                return '.java'
            if lang == 'cpp':
                return '.cpp'

        os.makedirs(dir_path, exist_ok=True)
        conn = sqlite3.connect(self.db_name)
        c = conn.cursor()
        c.execute(
            '''
                SELECT problem.id, submission.title 
                FROM submission 
                LEFT JOIN problem 
                    ON submission.title=problem.title 
                WHERE submission.downloaded=1 AND submission.removed=0 
                GROUP BY submission.title
            ''')
        titles = c.fetchall()
        for id, title in titles:
            if not id:
                continue
            problem_dir = os.path.join(dir_path, str(id).zfill(3) + '. ' + title)
            os.makedirs(problem_dir, exist_ok=True)
            c.execute('SELECT lang FROM submission WHERE downloaded=1 AND removed=0 AND title=?', (title,))
            langs = c.fetchall()
            langs = [lang[0] for lang in langs]
            for lang in langs:
                current_dir = os.path.join(problem_dir, lang_to_language(lang))
                os.makedirs(current_dir, exist_ok=True)
                c.execute(
                    'SELECT path FROM submission WHERE downloaded=1 AND removed=0 AND title=? AND lang=? ORDER BY url',
                    (title, lang))
                orig_file_paths = c.fetchall()
                orig_file_paths = [orig_file_path[0] for orig_file_path in orig_file_paths]
                shutil.copyfile(orig_file_paths[0], os.path.join(current_dir, 'Solution' + lang_to_extension(lang)))
                if latest_only:
                    continue
                for i in range(1, len(orig_file_paths)):
                    shutil.copyfile(orig_file_paths[0],
                                    os.path.join(current_dir, 'Solution ' + 'I' * (i + 1) + lang_to_extension(lang)))

        conn.close()

    def save_problem_list(self, file_name, file_type='csv', language='Chinese'):
        def preprocess(problem_list):
            for problem in problem_list:
                problem['acceptance'] = problem['total_acs'] / problem['total_submitted']
                problem['status'] = problem['status'] == 'ac'

        def to_locale(problem_list, language_dict):
            problem_list = [{language_dict[key]: value for (key, value) in problem.items()} for problem in problem_list]

            for problem in problem_list:
                problem[language_dict['difficulty']] = language_dict['level'][problem[language_dict['difficulty']]]
                problem[language_dict['paid_only']] = language_dict['bool'][problem[language_dict['paid_only']]]
                problem[language_dict['status']] = language_dict['bool'][problem[language_dict['status']]]
            return problem_list

        problem_list = self.get_problem_list_from_db()
        preprocess(problem_list)
        language_dict = self.get_language_dict(language)
        problem_list = to_locale(problem_list, language_dict)
        if file_type == 'csv':
            self.save_problem_list_as_csv(problem_list, file_name)
        elif file_type == 'excel':
            self.save_problem_list_as_excel(problem_list, file_name, language_dict)

    def save_problem_list_as_csv(self, problem_list, file_name):
        with open(file_name, 'w', encoding='utf-8', newline='') as f:
            import csv
            writer = csv.DictWriter(f, fieldnames=problem_list[0].keys())
            writer.writeheader()
            writer.writerows(problem_list)

    def save_problem_list_as_excel(self, problem_list, file_name, language_dict):
        from openpyxl import Workbook
        from openpyxl.styles import NamedStyle
        from openpyxl.formatting.rule import CellIsRule, DataBarRule
        from openpyxl.styles import PatternFill

        def format_cell_style(ws, language_dict):
            style_int = NamedStyle('int')
            style_int.number_format = '0'
            style_str = NamedStyle('str')
            style_str.number_format = '@'
            style_pcnt = NamedStyle('pcnt')
            style_pcnt.number_format = '0.0%'
            for cell in ws[column_index[language_dict['id']]][1:]:
                cell.style = style_int
            for cell in ws[column_index[language_dict['total_submitted']]][1:]:
                cell.style = style_int
            for cell in ws[column_index[language_dict['total_acs']]][1:]:
                cell.style = style_int
            for cell in ws[column_index[language_dict['title']]][1:]:
                cell.style = style_str
            for cell in ws[column_index[language_dict['slug']]][1:]:
                cell.style = style_str
            for cell in ws[column_index[language_dict['difficulty']]][1:]:
                cell.style = style_str
            for cell in ws[column_index[language_dict['paid_only']]][1:]:
                cell.style = style_str
            for cell in ws[column_index[language_dict['status']]][1:]:
                cell.style = style_str
            for cell in ws[column_index[language_dict['acceptance']]][1:]:
                cell.style = style_pcnt

        def conditional_formatting(ws, language_dict):
            def get_entire_column(index):
                return index + '1:' + index + '1048576'

            red_color = 'ffc7ce'
            green_color = 'c2efcf'
            yellow_color = 'ffeba2'

            red_fill = PatternFill(start_color=red_color, end_color=red_color, fill_type='solid')
            green_fill = PatternFill(start_color=green_color, end_color=green_color, fill_type='solid')
            yellow_fill = PatternFill(start_color=yellow_color, end_color=yellow_color, fill_type='solid')

            ws.conditional_formatting.add(get_entire_column(column_index[language_dict['difficulty']]),
                                          CellIsRule(operator='equal', formula=['"' + language_dict['level'][1] + '"'],
                                                     stopIfTrue=False, fill=green_fill))
            ws.conditional_formatting.add(get_entire_column(column_index[language_dict['difficulty']]),
                                          CellIsRule(operator='equal', formula=['"' + language_dict['level'][2] + '"'],
                                                     stopIfTrue=False, fill=yellow_fill))
            ws.conditional_formatting.add(get_entire_column(column_index[language_dict['difficulty']]),
                                          CellIsRule(operator='equal', formula=['"' + language_dict['level'][3] + '"'],
                                                     stopIfTrue=False, fill=red_fill))

            ws.conditional_formatting.add(get_entire_column(column_index[language_dict['paid_only']]),
                                          CellIsRule(operator='equal',
                                                     formula=['"' + language_dict['bool'][False] + '"'],
                                                     stopIfTrue=False, fill=green_fill))
            ws.conditional_formatting.add(get_entire_column(column_index[language_dict['paid_only']]),
                                          CellIsRule(operator='equal',
                                                     formula=['"' + language_dict['bool'][True] + '"'],
                                                     stopIfTrue=False, fill=red_fill))

            ws.conditional_formatting.add(get_entire_column(column_index[language_dict['status']]),
                                          CellIsRule(operator='equal',
                                                     formula=['"' + language_dict['bool'][False] + '"'],
                                                     stopIfTrue=False, fill=red_fill))
            ws.conditional_formatting.add(get_entire_column(column_index[language_dict['status']]),
                                          CellIsRule(operator='equal',
                                                     formula=['"' + language_dict['bool'][True] + '"'],
                                                     stopIfTrue=False, fill=green_fill))

            ws.conditional_formatting.add(get_entire_column(column_index[language_dict['acceptance']]),
                                          DataBarRule(start_type='percentile', start_value=0, end_type='percentile',
                                                      end_value=100, color="FF638EC6", showValue='None'))

        wb = Workbook()
        ws = wb.active
        ws.append(tuple(problem_list[0].keys()))
        column_index = {item.value: item.column for item in ws[1]}
        rows = [{column_index[key]: value for (key, value) in problem.items()} for problem in problem_list]
        for row in rows:
            ws.append(row)
        format_cell_style(ws, language_dict)
        conditional_formatting(ws, language_dict)
        wb.save(file_name)

    def get_language_dict(self, language):
        language_dict = None
        if language == 'Chinese':
            language_dict = {
                'id': '题号',
                'title': '标题',
                'slug': '链接',
                'difficulty': '难度',
                'total_submitted': '总提交数',
                'total_acs': '总通过数',
                'acceptance': '通过率',
                'paid_only': '付费',
                'status': '已解决',
                'level': {
                    1: '简单',
                    2: '中等',
                    3: '难'
                },
                'bool': {
                    True: '是',
                    False: '否'
                }
            }
        elif language == 'English':
            language_dict = {
                'id': '#',
                'title': 'Title',
                'slug': 'Link',
                'difficulty': 'Difficulty',
                'total_submitted': 'Total Submitted',
                'total_acs': 'Total Accepted',
                'acceptance': 'Acceptance',
                'paid_only': 'Paid Only',
                'status': 'Solved',
                'level': {
                    1: 'Easy',
                    2: 'Medium',
                    3: 'Hard'
                },
                'bool': {
                    True: 'Yes',
                    False: 'No'
                }
            }
        return language_dict


if __name__ == '__main__':
    extractor = Extractor()
    extractor.extract_descriptions()
